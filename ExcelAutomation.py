# Excel Automation Python
# Copy the excel file in the projects folder

import openpyxl as xl                                      # Import openpyxl module and give it an alias 'xl'
from openpyxl.chart import PieChart3D, Reference           # This will import charts

wb = xl.load_workbook('TransactionsAutomationPython.xlsx') # Make an object and use the load workbook function of openpyxl to load your excel sheet
sheet = wb['Sheet1']                                       # Make an object sheet to load the sheet from the loaded workbook

# cell = sheet['a1]                                        # THis will give you the value of the A1 cell in the sheet 1 of the workbook
# cell = sheet.cell(1, 1)                                  # Alternatively, you can also get the value of the A1 cell by mentioning the coordinates from the sheet
                                                            # print(sheet.max_row)  # this will print you the number of rows
for row in range(4, sheet.max_row + 1):                    # Make a for loop give the range and plus 1 because we need to show the last row also
    cell = sheet.cell(row, 5)                               # To get the access of this particular row
    new_price = cell.value * 1.8                            # To get the corrected price from accessed row
    new_price_cell = sheet.cell(row, 6)                 # this will make a new column for the new price
    new_price_cell.value = new_price                        # This object will store the value


values = Reference(sheet,                                   # Here we will assign which values from which row and columns we need to display in chart
                   min_row=4, max_row=sheet.max_row,
                   min_col=6,
                   max_col=6)

chart = PieChart3D()                                         # Assign instance 'pie_chart'
chart.add_data(values)                                      # Now Add data values containing the data
sheet.add_chart(chart, 'b10')                               # Add the coordinates where we need to display the chart


wb.save('TransactionsAutomationPythonUPDATED.xlsx')         # Save the updated file as a new file